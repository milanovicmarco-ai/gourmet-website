"""
Bulk import desde Excel DIRECTO a Neon Postgres.
Bypasea la API FastAPI: escribe directo en la tabla products.
~30 segundos para los 693 productos en vez de ~20 min vía API.

Uso (desde la raíz del repo):
  pip3 install psycopg2-binary openpyxl --break-system-packages
  export DATABASE_URL="postgresql://neondb_owner:npg_dXES3a4wgcsk@ep-still-mud-a49vkv9r.us-east-1.aws.neon.tech/neondb?sslmode=require"
  python3 scripts/bulk-import-neon.py /Users/marco/MyProjects/Productos/aurellano-catalogo-optimizado.xlsx

Política:
  - Excel = fuente de verdad
  - Celda vacía → NO toca el valor existente (preserva)
  - Celda con valor "basura" (sentinel "-", ".", etc. que normaliza a string vacío) → NULL en DB
  - Transacción: si algo falla, rollback completo. Cero riesgo.
"""

import os
import sys
import re
from decimal import Decimal
from datetime import date
from typing import Optional
import unicodedata

try:
    import psycopg2
    import psycopg2.extras
    import openpyxl
except ImportError:
    print("Falta instalar dependencias. Ejecuta:")
    print("  pip3 install psycopg2-binary openpyxl --break-system-packages")
    sys.exit(1)


# Map header Excel → columna DB
HEADER_MAP = {
    "ref": "ref",
    "nombre": "name",
    "familia": "family",
    "estado": "status",
    "marca": "brand",
    "descripcion_corta": "descripcion_corta",
    "descripcion_larga": "description_rich",
    "origen": "origen",
    "sabor": "flavor",
    "unidades_por_caja": "units_per_box",
    "precio_eur": "base_price_eur",
    "alergenos": "alergenos",
    "ingredientes": "ingredientes",
    "seo_title": "seo_title",
    "seo_description": "seo_description",
    "sin_gluten": "sin_gluten",
    "sin_lactosa": "sin_lactosa",
    "vegetariano": "vegetariano",
}

INT_COLS = {"units_per_box"}
DECIMAL_COLS = {"base_price_eur"}
BOOL_COLS = {"sin_gluten", "sin_lactosa", "vegetariano"}


def is_junk(s: str) -> bool:
    """Detecta sentinels tipo '-', '—', '·', '.', 'n/a' — strings que tras
    quitar acentos y no alfanuméricos quedan vacíos."""
    if s is None:
        return True
    norm = unicodedata.normalize("NFD", str(s).lower())
    norm = "".join(c for c in norm if unicodedata.category(c) != "Mn")
    norm = re.sub(r"[^a-z0-9]+", "", norm).strip()
    return len(norm) == 0


def parse_bool(v) -> bool:
    if isinstance(v, bool):
        return v
    return str(v).strip().lower() in ("true", "1", "si", "sí", "yes", "y", "x")


def parse_status(v) -> Optional[str]:
    s = str(v or "").strip().lower()
    if s in ("draft", "borrador"):
        return "draft"
    if s in ("published", "publicado", "publish"):
        return "published"
    if s in ("archived", "archivado", "archive"):
        return "archived"
    return None


def normalize_value(db_col: str, val):
    """Convierte el valor del Excel al tipo esperado por la DB.
    Devuelve (valor_normalizado, debe_aplicarse).
    debe_aplicarse=False si el valor es vacío y NO debe tocarse en DB."""
    if val is None:
        return None, False
    if isinstance(val, str) and val.strip() == "":
        return None, False

    if db_col == "status":
        st = parse_status(val)
        if st is None:
            return None, False
        return st, True

    if db_col in BOOL_COLS:
        return parse_bool(val), True

    if db_col in INT_COLS:
        try:
            return int(float(val)), True
        except (ValueError, TypeError):
            return None, False

    if db_col in DECIMAL_COLS:
        try:
            return Decimal(str(val).replace(",", ".")), True
        except Exception:
            return None, False

    # String fields: si es "basura" (sentinel "-" o similar), poner NULL
    if isinstance(val, str):
        s = val.strip()
        if is_junk(s):
            return None, True  # explicitly set NULL
        return s, True

    return val, True


def slugify(name: str) -> str:
    """Convierte un nombre en slug URL-safe."""
    s = unicodedata.normalize("NFD", name.lower())
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s


def resolve_brand(cur, brand_text, cache):
    """Resuelve nombre de marca → slug. Crea la marca si no existe.
    Devuelve None si brand_text es vacío o basura."""
    if brand_text is None:
        return None
    s = str(brand_text).strip()
    if not s or is_junk(s):
        return None
    key = s.lower()
    if key in cache:
        return cache[key]
    slug = slugify(s)
    if not slug:
        cache[key] = None
        return None
    # ¿Ya existe?
    cur.execute(
        "SELECT slug FROM brands WHERE LOWER(name) = LOWER(%s) OR slug = %s LIMIT 1",
        (s, slug),
    )
    row = cur.fetchone()
    if row:
        cache[key] = row[0]
        return row[0]
    # Insertar nueva
    cur.execute(
        "INSERT INTO brands (slug, name) VALUES (%s, %s) ON CONFLICT (slug) DO UPDATE SET name=EXCLUDED.name RETURNING slug",
        (slug, s),
    )
    new_slug = cur.fetchone()[0]
    cache[key] = new_slug
    return new_slug


def resolve_family(cur, fam_text, cache):
    """Igual que resolve_brand pero para families. El slug de familia suele
    ser MAYÚSCULAS_CON_GUIONES (VARIOS, FOIE_GRAS, etc.)."""
    if fam_text is None:
        return None
    s = str(fam_text).strip()
    if not s or is_junk(s):
        return None
    key = s.lower()
    if key in cache:
        return cache[key]
    # Si ya parece un slug (todo mayúsculas/números/guiones), úsalo
    if re.fullmatch(r"[A-Z0-9_]+", s):
        slug = s
    else:
        slug = re.sub(r"[^A-Z0-9]+", "_",
                      unicodedata.normalize("NFD", s.upper())
                      .encode("ascii", "ignore").decode()).strip("_")
    if not slug:
        cache[key] = None
        return None
    cur.execute(
        "SELECT slug FROM families WHERE LOWER(name) = LOWER(%s) OR slug = %s LIMIT 1",
        (s, slug),
    )
    row = cur.fetchone()
    if row:
        cache[key] = row[0]
        return row[0]
    cur.execute(
        "INSERT INTO families (slug, name) VALUES (%s, %s) ON CONFLICT (slug) DO UPDATE SET name=EXCLUDED.name RETURNING slug",
        (slug, s),
    )
    new_slug = cur.fetchone()[0]
    cache[key] = new_slug
    return new_slug


def load_rows(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row):
            continue
        rec = dict(zip(headers, row))
        rows.append(rec)
    return rows


def main():
    if len(sys.argv) < 2:
        print("Uso: python3 scripts/bulk-import-neon.py <ruta/excel.xlsx>")
        sys.exit(1)

    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        print("Falta env var DATABASE_URL. Configúrala con:")
        print('  export DATABASE_URL="postgresql://..."')
        sys.exit(1)

    excel_path = sys.argv[1]
    print(f"Excel:    {excel_path}")
    print(f"Neon DB:  {db_url.split('@')[1].split('/')[0]}")
    print()

    print("Leyendo Excel...")
    rows = load_rows(excel_path)
    print(f"  {len(rows)} filas con datos")
    print()

    print("Conectando a Neon...")
    conn = psycopg2.connect(db_url, connect_timeout=15)
    conn.autocommit = False
    print("  ✓ Conectado")

    try:
        with conn.cursor() as cur:
            cur.execute("SELECT count(*) FROM products")
            before = cur.fetchone()[0]
            print(f"  Total productos en DB ANTES: {before}")
            print()

            updated = 0
            inserted = 0
            unchanged = 0
            errors = []
            today = date.today()
            brand_cache = {}
            family_cache = {}

            for i, rec in enumerate(rows):
                line = i + 2
                ref_raw = rec.get("ref")
                ref = str(ref_raw).strip() if ref_raw is not None else ""
                if not ref:
                    errors.append(f"Fila {line}: sin ref, ignorada")
                    continue

                # Construir el dict de campos a aplicar
                payload = {}
                for excel_col, db_col in HEADER_MAP.items():
                    if excel_col == "ref":
                        continue  # ref es PK, no se actualiza
                    if excel_col not in rec:
                        continue
                    val_norm, apply = normalize_value(db_col, rec[excel_col])
                    if apply:
                        payload[db_col] = val_norm

                # Resolver brand: name → slug (crea si no existe)
                if "brand" in payload:
                    payload["brand"] = resolve_brand(cur, payload["brand"], brand_cache)
                # Resolver family: name → slug
                if "family" in payload:
                    payload["family"] = resolve_family(cur, payload["family"], family_cache)

                # ¿Existe ya el producto?
                cur.execute("SELECT 1 FROM products WHERE ref = %s", (ref,))
                exists = cur.fetchone() is not None

                if exists:
                    if not payload:
                        unchanged += 1
                        continue
                    # Sincronizar active si tocamos status
                    if "status" in payload:
                        payload["active"] = payload["status"] != "archived"
                    payload["fecha_modificacion"] = today
                    set_clause = ", ".join(f"{k} = %s" for k in payload.keys())
                    cur.execute(
                        f"UPDATE products SET {set_clause} WHERE ref = %s",
                        list(payload.values()) + [ref],
                    )
                    updated += 1
                else:
                    payload["ref"] = ref
                    payload["fecha_modificacion"] = today
                    if "status" in payload:
                        payload["active"] = payload["status"] != "archived"
                    cols = list(payload.keys())
                    placeholders = ", ".join(["%s"] * len(cols))
                    cur.execute(
                        f"INSERT INTO products ({', '.join(cols)}) VALUES ({placeholders})",
                        list(payload.values()),
                    )
                    inserted += 1

                if (i + 1) % 100 == 0:
                    print(f"  Procesadas {i+1}/{len(rows)}...")

            cur.execute("SELECT count(*) FROM products")
            after = cur.fetchone()[0]

        print()
        print("Aplicando transacción...")
        conn.commit()
        print("  ✓ COMMIT")
        print()
        print("─" * 60)
        print(f"Resumen:")
        print(f"  Actualizados: {updated}")
        print(f"  Creados:      {inserted}")
        print(f"  Sin cambios:  {unchanged}")
        print(f"  Errores:      {len(errors)}")
        print(f"  Total en DB ANTES:  {before}")
        print(f"  Total en DB DESPUÉS: {after}")
        if errors:
            print()
            print("Errores:")
            for e in errors[:30]:
                print(f"  {e}")
            if len(errors) > 30:
                print(f"  ... +{len(errors)-30} más")

    except Exception as e:
        conn.rollback()
        print()
        print(f"✖ ERROR FATAL: {e}")
        print("  → ROLLBACK aplicado, la BD NO se ha tocado.")
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    main()
