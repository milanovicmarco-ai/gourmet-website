"""
Bulk upload de traducciones (CA o cualquier locale) desde un Excel
DIRECTO a Supabase (tabla product_translations).

Política:
  - Excel = fuente de verdad para los campos textuales traducibles.
  - Celda vacía → NO se toca el valor existente (preserva).
  - Celda con valor "basura" (sentinel "-", ".", etc.) → NULL en la DB,
    de forma que el detalle público hace fallback al ES canónico.
  - Solo upsertea filas con ref presente en el Excel; no borra nada.
  - Transacción: si algo falla, rollback completo. Cero riesgo.

Uso (desde la raíz del repo):
  pip3 install psycopg2-binary openpyxl     # si no los tienes
  export SUPABASE_DB_URL="postgresql://postgres.<ref>:<pwd>@aws-0-eu-west-1.pooler.supabase.com:6543/postgres?sslmode=require"
  python3 scripts/bulk-upload-translations.py \\
      /ruta/a/aurellano-catalogo-CA.xlsx \\
      --locale ca

  # Para dry-run (mira qué pasaría sin commitear):
  python3 scripts/bulk-upload-translations.py /ruta/excel.xlsx --locale ca --dry-run

¿Dónde sacar la SUPABASE_DB_URL?
  Supabase dashboard → Project Settings → Database → Connection string →
  "URI" en modo "Transaction pooler" (puerto 6543). Copia, sustituye
  [YOUR-PASSWORD] por el real. Esa cadena es la que va en SUPABASE_DB_URL.
"""

import os
import re
import sys
import argparse
import unicodedata
from datetime import datetime, timezone

try:
    import psycopg2
    import psycopg2.extras
    import openpyxl
except ImportError:
    print("Falta instalar dependencias. Ejecuta:")
    print("  pip3 install psycopg2-binary openpyxl")
    sys.exit(1)


# Campos traducibles aceptados. Cualquier otra columna del Excel se ignora.
# El header del Excel se busca en case-insensitive: 'NAME', 'Name', 'name' valen.
TRANSLATABLE_COLS = {
    "name",
    "descripcion_corta",
    "description_rich",
    "flavor",
    "origen",
    "ingredientes",
    "seo_title",
    "seo_description",
}

# El export del PIM usa headers en castellano para 3 columnas cuyo nombre en
# la DB es la versión inglesa. Los mapeamos automáticamente para que el Excel
# CA siga teniendo las cabeceras del export sin que el usuario tenga que
# renombrar nada a mano.
HEADER_ALIASES = {
    "nombre": "name",
    "descripcion_larga": "description_rich",
    "sabor": "flavor",
}


def is_junk(s) -> bool:
    """Detecta sentinels tipo '-', '—', '·', '.', 'n/a' — strings que
    tras quitar acentos y no alfanuméricos quedan vacíos."""
    if s is None:
        return True
    norm = unicodedata.normalize("NFD", str(s).lower())
    norm = "".join(c for c in norm if unicodedata.category(c) != "Mn")
    norm = re.sub(r"[^a-z0-9]+", "", norm).strip()
    return len(norm) == 0


def load_rows(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    raw_headers = [
        str(c.value).strip().lower() if c.value is not None else ""
        for c in ws[1]
    ]
    # Aplica alias: 'nombre' → 'name', 'sabor' → 'flavor', etc.
    headers = [HEADER_ALIASES.get(h, h) for h in raw_headers]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row):
            continue
        rows.append(dict(zip(headers, row)))
    return headers, rows


def main():
    parser = argparse.ArgumentParser(description="Bulk upload de traducciones a Supabase.")
    parser.add_argument("excel", help="Ruta al Excel con la traducción.")
    parser.add_argument(
        "--locale",
        default="ca",
        help="Locale destino (default: 'ca'). Cualquier valor admitido por la columna locale.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="No commitea. Muestra qué pasaría.",
    )
    args = parser.parse_args()

    db_url = os.environ.get("SUPABASE_DB_URL")
    if not db_url:
        print("Falta env var SUPABASE_DB_URL.")
        print("  export SUPABASE_DB_URL=\"postgresql://postgres.<ref>:<pwd>@aws-0-eu-west-1.pooler.supabase.com:6543/postgres?sslmode=require\"")
        sys.exit(1)

    print(f"Excel:    {args.excel}")
    print(f"Supabase: {db_url.split('@')[1].split('/')[0]}")
    print(f"Locale:   {args.locale}")
    print(f"Dry-run:  {args.dry_run}")
    print()

    print("Leyendo Excel...")
    headers, rows = load_rows(args.excel)
    print(f"  {len(rows)} filas con datos.")
    translatable_in_excel = [h for h in headers if h in TRANSLATABLE_COLS]
    ignored = [h for h in headers if h and h not in TRANSLATABLE_COLS and h != "ref"]
    print(f"  Columnas traducibles encontradas: {translatable_in_excel}")
    if ignored:
        print(f"  Columnas ignoradas (no traducibles o no en schema): {ignored}")
    if "ref" not in headers:
        print("  ✖ El Excel NO tiene columna 'ref'. Imposible identificar productos.")
        sys.exit(1)
    print()

    print("Conectando a Supabase...")
    conn = psycopg2.connect(db_url, connect_timeout=15)
    conn.autocommit = False
    print("  ✓ Conectado")

    inserted = 0
    updated = 0
    skipped_empty = 0
    skipped_no_ref = 0
    errors = []
    now = datetime.now(timezone.utc)

    try:
        with conn.cursor() as cur:
            # Sanity: ¿existe la tabla?
            cur.execute(
                "SELECT count(*) FROM product_translations WHERE locale = %s",
                (args.locale,),
            )
            before = cur.fetchone()[0]
            print(f"  Traducciones en '{args.locale}' ANTES: {before}")
            print()

            for i, rec in enumerate(rows):
                line = i + 2  # +1 por header, +1 por 1-indexed
                ref_raw = rec.get("ref")
                ref = str(ref_raw).strip() if ref_raw is not None else ""
                if not ref:
                    skipped_no_ref += 1
                    continue

                # Construir payload solo con campos traducibles que estén en el Excel.
                payload = {}
                for col in TRANSLATABLE_COLS:
                    if col not in rec:
                        continue
                    val = rec[col]
                    if val is None:
                        continue  # celda vacía → preservar existente
                    if isinstance(val, str):
                        s = val.strip()
                        if s == "":
                            continue  # celda vacía → preservar
                        if is_junk(s):
                            payload[col] = None  # sentinel → NULL explícito
                        else:
                            payload[col] = s
                    else:
                        # Numero, etc.: castea a string por si acaso.
                        payload[col] = str(val).strip()

                if not payload:
                    skipped_empty += 1
                    continue

                # ¿Ya existe traducción para este ref+locale?
                cur.execute(
                    "SELECT 1 FROM product_translations WHERE product_ref = %s AND locale = %s",
                    (ref, args.locale),
                )
                exists = cur.fetchone() is not None

                payload["updated_at"] = now
                if exists:
                    set_clause = ", ".join(f"{k} = %s" for k in payload.keys())
                    try:
                        cur.execute(
                            f"UPDATE product_translations SET {set_clause} "
                            "WHERE product_ref = %s AND locale = %s",
                            list(payload.values()) + [ref, args.locale],
                        )
                        updated += 1
                    except Exception as e:
                        errors.append(f"Fila {line} (ref {ref}): UPDATE falló: {e}")
                else:
                    payload["product_ref"] = ref
                    payload["locale"] = args.locale
                    cols = list(payload.keys())
                    placeholders = ", ".join(["%s"] * len(cols))
                    try:
                        cur.execute(
                            f"INSERT INTO product_translations ({', '.join(cols)}) "
                            f"VALUES ({placeholders})",
                            list(payload.values()),
                        )
                        inserted += 1
                    except Exception as e:
                        errors.append(f"Fila {line} (ref {ref}): INSERT falló: {e}")

                if (i + 1) % 100 == 0:
                    print(f"  Procesadas {i+1}/{len(rows)}...")

            cur.execute(
                "SELECT count(*) FROM product_translations WHERE locale = %s",
                (args.locale,),
            )
            after = cur.fetchone()[0]

        if args.dry_run:
            conn.rollback()
            print()
            print("  → DRY-RUN: ROLLBACK aplicado, nada se ha guardado.")
        else:
            print()
            print("Aplicando transacción...")
            conn.commit()
            print("  ✓ COMMIT")

        print()
        print("─" * 60)
        print("Resumen:")
        print(f"  Insertadas:           {inserted}")
        print(f"  Actualizadas:         {updated}")
        print(f"  Sin cambios (vacías): {skipped_empty}")
        print(f"  Sin ref (ignoradas):  {skipped_no_ref}")
        print(f"  Errores:              {len(errors)}")
        print(f"  Total en '{args.locale}' ANTES:   {before}")
        print(f"  Total en '{args.locale}' DESPUÉS: {after}")
        if errors:
            print()
            print("Errores:")
            for e in errors[:30]:
                print(f"  {e}")
            if len(errors) > 30:
                print(f"  ... +{len(errors)-30} más")

    except Exception as e:
        conn.rollback()
        print()
        print(f"✖ ERROR FATAL: {e}")
        print("  → ROLLBACK aplicado, la BD NO se ha tocado.")
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    main()
